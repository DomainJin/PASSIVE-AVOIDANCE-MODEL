import tkinter as tk
from tkinter import filedialog
import openpyxl
from datetime import datetime

def get_second_last_value(file_path):
    time_change = 0
    if file_path:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        # Tìm ô cuối cùng có dữ liệu trong cột A
        last_row = ws.max_row
        
        # Nếu có ít nhất 2 hàng, lấy ô kế cuối
        if last_row >= 2:
            last_cell = ws['A' + str(last_row - 1)]
            last_value = last_cell.value
            second_last_cell = ws['A' + str(last_row - 2)]
            second_last_value = second_last_cell.value
           
            if isinstance(last_value, datetime) and isinstance(second_last_value, datetime):
                last_value_float = last_value.timestamp()
                second_last_value_float = second_last_value.timestamp()
                time_change = last_value_float - second_last_value_float
        else:
            print("File Excel không đủ dữ liệu để lấy ô kế cuối")
    return time_change

def select_file():
    root = tk.Tk()
    root.withdraw()  # Ẩn cửa sổ chính
    file_path = filedialog.askopenfilename(
        title="Chọn tệp để mở",
        filetypes=[("Excel files", "*.xlsx")]
    )
    return file_path
def write_result_to_excel(file_path, result):
    """Ghi kết quả vào ô B2 của file Excel"""
    try:
        # Mở file để ghi (không ở chế độ read_only)
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Ghi giá trị vào ô D2
        ws['D1'] = 'Latency Time'
        ws['D2'] = result
        
        # Lưu lại file
        wb.save(file_path)
        print(f"Đã lưu kết quả {result} vào ô D2 của file {file_path}")
        return True
    except Exception as e:
        print(f"Lỗi khi ghi vào file Excel: {e}")
        return False
# Ví dụ sử dụng:
if __name__ == "__main__":
    file_path = select_file()
    value = get_second_last_value(file_path)
    
    if file_path:
        file2 = select_file()
        value2 = get_second_last_value(file2)
        result = value + value2
        print(f"Value: {result}")
        write_result_to_excel(file_path, result)
    else:
        print("Không có tệp nào được chọn.")