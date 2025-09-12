import cv2
from picamera2 import Picamera2
from ultralytics import YOLO
from ultralytics import solutions
from main1 import Ui_MainWindow
from PySide6.QtWidgets import QApplication, QMainWindow, QLabel, QVBoxLayout, QPushButton, QTableWidget,QTableWidgetItem,QWidget
from PySide6.QtGui import QImage, QPixmap
from PySide6.QtCore import QTimer
import pandas as pd
import resource_rc
from PySide6.QtCore import Qt
import minimalmodbus
import serial
import time
import sys
import os
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import numpy as np
from PIL import Image
import matplotlib.pyplot as plt

class MyServer(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle("Passive Avoidance Test")
        self.instrument = minimalmodbus.Instrument('/dev/ttyUSB0', 10)  # port name, slave address (in decimal)
        self.instrument.serial.timeout = 3
        self.U = self.instrument.read_register(0, 0)/100  # Registernumber, number of decimals
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.load_mearsure)  # Khi h?t th?i gian, g?i load_mearsure
        self.timer.start(1000) 
        # ledstatus = input("Give in 0 or 1 to turn on or off the led: ")
        # instrument.write_register(2, int(ledstatus), 0) 
        # Initialize the Picamera2
        self.picam2 = Picamera2()
        self.picam2.preview_configuration.main.size = (1280, 700)
        self.picam2.preview_configuration.main.format = "RGB888"
        self.picam2.preview_configuration.align()
        self.picam2.configure("preview")
        self.picam2.start()
        # Load the YOLO model
        self.model = YOLO("best1.pt")
        # Initialize heatmap generation using the ultralytics solution
        self.heatmap = solutions.Heatmap(colormap=cv2.COLORMAP_PARULA, show=True, model="best1.pt", classes=[1])
        # Setup the tracking QLabel for displaying the frame
        self.tracking = QLabel(self.grouptrack)  # Assuming self.grouptrack is your layout/container
        self.tracking.setGeometry(60, 30, 507, 245)  # Set to the desired size
        # Setup the heatmap QLabel for displaying the heatmap
        self.heatmap_label = QLabel(self.groupheatmap)  # Assuming self.groupheatmap is your layout/container
        self.heatmap_label.setGeometry(60,30, 507, 245)  # Set to the desired size
        # 
        # Layout and widget
        # Create a QTableWidget
        self.excel = QTableWidget(self.group_excel)
        # self.gridLayout_6.addWidget(self.excel, 0, 0, 1, 1)
        #
        self.chart = QLabel(self.group_chart)
        self.chart.setGeometry(0,30, 507, 245)  
        self.group_main.setVisible(False)
        self.group_output.setVisible(False)
        self.resize(1208, 300)  # ï¿½i?u ch?nh kï¿½ch thu?c c?a s? sao cho v?a d? (thay d?i chi?u cao cho phï¿½ h?p)
        # Optional: Create a timer for continuous frame capture (initially stopped)
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_camera_feed)
        # Connect button_start to start camera feed
        self.button_start.clicked.connect(self.on_button_start_clicked)
        self.previous_state = "Start"
        self.state_str = "Start"
        # B? d?m th?i gian (timer) d? x? lï¿½ 5 giï¿½y
        self.right_border_timer = None

        # Luu tr?ng thï¿½i c?a chu?t d? trï¿½nh th?c hi?n l?i hï¿½nh d?ng
        self.mouse_in_right_position = False
    def load_mearsure(self):
        try:
            # ï¿½?c giï¿½ tr? t? Modbus
            self.U = self.instrument.read_register(0, 0) / 100  # ï¿½i?n ï¿½p
            self.I = self.instrument.read_register(1, 0) / 10   # Dï¿½ng di?n
            # Hi?n th? cï¿½c giï¿½ tr? lï¿½n QLCDNumber
            self.volt.display(self.U)
            self.amp.display(self.I)
            # In ra giï¿½ tr? d? debug
            # print(f"C?p nh?t giï¿½ tr?: ï¿½i?n ï¿½p = {self.U}, Dï¿½ng di?n = {self.I}")
        except Exception as e:
            print(f"Doc ko dc {e}")
    def on_button_start_clicked(self):
        # When the button is clicked, start capturing frames and stop automatic updates
        id = self.idmouse.currentText()
        st = self.status.currentText()
        dr = self.drug.currentText()
        runtime = self.runtime.value()  # L?y giï¿½ tr? runtime t? QSpinBox
        # T?o thu m?c luu tr?
        global folder
        folder = self.create_folder(id, st, dr)
        global excel_file
        excel_file = os.path.join(folder, "data.xlsx")
        print(excel_file)
        # self.load_excel_file(excel_file)
        # Show the group_output after the button is clicked
        self.group_output.setVisible(True)
        self.group_excel.setVisible(True)
        self.timer.start(30)  # Start the timer with 30 ms interval
        self.update_camera_feed()  # Also call update_camera_feed immediately      
        # Show the group_output after the button is clicked
        self.group_output.setVisible(True)
        self.previous_time = datetime.now()
    def update_camera_feed(self):
        self.group_main.setVisible(True)
        self.group_output.setVisible(True)
        
        # Thay d?i kï¿½ch thu?c c?a s? sao cho phï¿½ h?p v?i n?i dung
        self.resize(1208, 882)
        # Capture frame-by-frame
        frame = self.picam2.capture_array()

        # Run YOLO inference on the frame
        results = self.model(frame)

        # Extract boxes and class information
        boxes = results[0].boxes
        class_ids = boxes.cls  # Get class IDs for each box

        # Generate the heatmap for the current frame
        im0_with_heatmap = self.heatmap.generate_heatmap(frame)
        image_path = os.path.join(folder, 'heatmap_image.jpg')

# Save the image
        cv2.imwrite(image_path, im0_with_heatmap)
        # Initialize variables for the border and lab-mouse
        border = None
        lab_mouse = None
        
        for i, box in enumerate(boxes):
            class_id = class_ids[i]  # Get the class ID for each box
            
            if class_id == 0:  # Assuming 'border' class is labeled as 0
                border = box.xyxy[0]  # [x1, y1, x2, y2] format for the border
            elif class_id == 1:  # Assuming 'lab-mouse' class is labeled as 1
                lab_mouse = box.xyxy[0]  # [x1, y1, x2, y2] format for the lab-mouse

        # Check if both border and lab-mouse are detected
        if border is not None and lab_mouse is not None:
            x1_border, y1_border, x2_border, y2_border = border
            x1_mouse, y1_mouse, x2_mouse, y2_mouse = lab_mouse
            current_time = time.time()
            if x2_mouse < x1_border:  # Lab-mouse is to the left of the border
                print("Lab-mouse is to the left of the border")
                self.state_str = "Black box to White box"
                self.instrument.write_register(2, 0, 0)
                self.instrument.write_register(3, 0, 0)
            elif x1_mouse > x2_border:  # Lab-mouse is to the right of the border
                print("Lab-mouse is to the right of the border")
                self.state_str = "White box to Black box"
                self.instrument.write_register(3, 1, 0)
                # Start or reset the timer if the mouse moves to the right
                if not self.mouse_in_right_position:
                    self.mouse_in_right_position = True
                    self.start_right_border_timer()

        if self.state_str != self.previous_state:
            self.save_change(self.previous_state)
            self.previous_state = self.state_str
            
        # Visualize or handle the rest of the processingsss
        annotated_frame = results[0].plot()

        # Convert OpenCV frame (BGR) to QImage (RGB)
        rgb_frame = cv2.cvtColor(annotated_frame, cv2.COLOR_BGR2RGB)
        height, width, channel = rgb_frame.shape
        bytes_per_line = 3 * width
        q_image = QImage(rgb_frame.data, width, height, bytes_per_line, QImage.Format_RGB888)
        
        # Convert heatmap (BGR) to QImage (RGB)
        rgb_frame1 = cv2.cvtColor(im0_with_heatmap, cv2.COLOR_BGR2RGB)
        height1, width1, channel1 = rgb_frame1.shape
        bytes_per_line1 = 3 * width1
        q_heatmap = QImage(rgb_frame1.data, width1, height1, bytes_per_line1, QImage.Format_RGB888)

        # Resize the frames to fit within 507x245 while maintaining aspect ratio
        new_width, new_height = self.calculate_aspect_ratio_resize(width, height, 507, 245)
        resized_image = q_image.scaled(new_width, new_height)
        resized_heatmap = q_heatmap.scaled(new_width, new_height)

        # Update QLabel with the resized image
        self.tracking.setPixmap(QPixmap.fromImage(resized_image))
        self.heatmap_label.setPixmap(QPixmap.fromImage(resized_heatmap))
        self.load_excel_file()
        total_time_on,total_time_off = self.total_time(excel_file)
        print(f"Total Time White: {total_time_on} second")
        print(f"Total Time Black: {total_time_off} second")
        self.create_chart(total_time_on,total_time_off)
        self.load_image(folder)
    def calculate_aspect_ratio_resize(self, original_width, original_height, target_width, target_height):
        # Calculate aspect ratios
        aspect_ratio = original_width / original_height
        target_aspect_ratio = target_width / target_height

        if aspect_ratio > target_aspect_ratio:
            # If the image is wider than the target size, adjust width
            new_width = target_width
            new_height = int(target_width / aspect_ratio)
        else:
            # If the image is taller than the target size, adjust height
            new_height = target_height
            new_width = int(target_height * aspect_ratio)
        return new_width, new_height
    def start_right_border_timer(self):
        """Starts a timer that waits 5 seconds before writing to the register."""
        if self.right_border_timer is None:
            self.right_border_timer = QTimer(self)
            self.right_border_timer.timeout.connect(self.on_right_border_timeout)
            self.right_border_timer.start(5000)  # 5 seconds delay

    def stop_right_border_timer(self):
        """Stops the 5-second timer if the mouse moves away from the right border."""
        if self.right_border_timer is not None:
            self.right_border_timer.stop()
            self.right_border_timer = None
            self.mouse_in_right_position = False  # Reset the flag

    def on_right_border_timeout(self):
        """This function is called when the 5-second timer expires."""
        print("5 seconds elapsed. Writing to the register.")
        self.instrument.write_register(2, 1, 0)
        self.instrument.write_register(3, 0, 0)
        self.stop_right_border_timer()  # Stop the timer after it expires
    #display excel
    
    def load_excel_file(self):
    # Ki?m tra n?u thu m?c vï¿½ t?p t?n t?i
        if os.path.exists(folder):  # Ki?m tra xem thu m?c cï¿½ t?n t?i khï¿½ng
            excel_file_path = os.path.join(folder, "data.xlsx")
            if os.path.exists(excel_file_path):  # N?u t?p t?n t?i
                df = pd.read_excel(excel_file_path)

                # ï¿½?t s? dï¿½ng vï¿½ s? c?t cho b?ng
                self.excel.setRowCount(df.shape[0])
                self.excel.setColumnCount(df.shape[1])

                # ï¿½?t tiï¿½u d? cho cï¿½c c?t
                self.excel.setHorizontalHeaderLabels(df.columns)

                # ï¿½i?n d? li?u vï¿½o b?ng t? DataFrame
                for row in range(df.shape[0]):
                    for col in range(df.shape[1]):
                        self.excel.setItem(row, col, QTableWidgetItem(str(df.iloc[row, col])))
            else:
                print(f"T?p {excel_file_path} khï¿½ng t?n t?i.")
        else:
            print(f"Thu m?c {folder} khï¿½ng t?n t?i.")
    def create_chart(self,x,y):
        left = [1, 2]
        height = [x,y]

            # labels for bars
        tick_label = ['White box', 'Black box']

                # plotting a bar chart
        plt.bar(left, height, tick_label = tick_label, width = 0.8, color = ['red', 'green'])
        plt.title('Chart of Position over Time')
        plt.xlabel('Position')
        plt.ylabel('Time (second)')
        plt.grid()
        plt.savefig(os.path.join(folder, 'chart.png'))  # Save the chart as a PNG file

    def load_image(self,folder):
        image = os.path.join(folder, 'chart.png')
        pixmap = QPixmap(image)
        scaled_pixmap = pixmap.scaled(511, 188, Qt.AspectRatioMode.KeepAspectRatio)
        self.chart.setPixmap(scaled_pixmap)

    def create_folder(self, x, y, z):
        current_time = datetime.now()

            # Format the date and time in YYYY-MM-DD HH:MM:SS
        formatted_time = current_time.strftime("%Y-%m-%d %H:%M:%S")
        folder_name = f"{formatted_time}_{x}_{y}_{z}"
        os.makedirs(folder_name, exist_ok=True)
        return folder_name

    def save_change(self,state):
        try:
            wb = openpyxl.load_workbook(excel_file)
        except FileNotFoundError:
            wb = Workbook()
        # L?y th?i gian hi?n t?i
        current_time = datetime.now()

        # ï¿½?nh d?ng th?i gian theo ki?u 'yyyy/mm/dd hh:mm:ss'
        formatted_time = current_time.strftime("%Y/%m/%d %H:%M:%S")
        sheet = wb.active
        max_row = sheet.max_row

        for row in range(max_row, 0, -1):
            sheet.cell(row=row + 1, column=1).value = sheet.cell(row=row, column=1).value
            sheet.cell(row=row + 1, column=2).value = sheet.cell(row=row, column=2).value

        sheet.cell(row=1, column=1).value = formatted_time
        sheet.cell(row=1, column=2).value = state
        wb.save(excel_file)
    def total_time(self,file):
        try:
            wb = openpyxl.load_workbook(file)
        except FileNotFoundError:
            print(f"File {file} not found.")
            return
        sheet = wb.active
        total_rows = sheet.max_row
        total_1 = 0
        total_2 = 0
        for i in range (total_rows,2,-2):
            raw1 = sheet.cell(row=i, column=1).value
            raw2 = sheet.cell(row=i-1, column=1).value
            raw3 = sheet.cell(row=i-2, column=1).value
            time1 = datetime.strptime(raw1, '%Y/%m/%d %H:%M:%S')
            time2 = datetime.strptime(raw2, '%Y/%m/%d %H:%M:%S')
            time3 = datetime.strptime(raw3, '%Y/%m/%d %H:%M:%S')
            total_1 += int(time1.timestamp()) - int(time2.timestamp())
            total_2 += int(time2.timestamp()) - int(time3.timestamp())
        print(-total_1, -total_2)
        return -total_1, -total_2
    def your_slot_function(self):
        # Your implementation here
        pass
# Run the application
if __name__ == "__main__":
    app = QApplication([])
    window = MyServer()
    window.show()
    app.exec()